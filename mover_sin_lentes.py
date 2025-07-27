#!/usr/bin/env python3
import os
import re
import shutil
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

excel_path = r"C:\Users\Administrador\Downloads\Reporte_2.xlsx"
target_dir = os.path.join('results', 'sin_lentes')
os.makedirs(target_dir, exist_ok=True)

# 1) Libro con valores (para Probabilidad)
wb_vals = load_workbook(excel_path, data_only=True)
ws_vals: Worksheet = wb_vals.worksheets[0]

# 2) Libro con fórmulas (para fórmula + objeto hyperlink)
wb_form = load_workbook(excel_path, data_only=False)
ws_form: Worksheet = wb_form.worksheets[0]

# Regex: HYPERLINK, _xlfn.HYPERLINK o HIPERVINCULO + primera cadena entre comillas
pat = re.compile(r'(?:_xlfn\.)?(?:HYPERLINK|HIPERVINCULO)\("([^"]+)"', re.IGNORECASE)

for (cel_val_ruta, cel_val_prob, _), (cel_form_ruta,) in zip(
        ws_vals.iter_rows(min_row=2, max_col=3),
        ws_form.iter_rows(min_row=2, max_col=1)):

    if cel_val_prob.value != 0:
        continue

    # 1️⃣  ¿Hay hyperlink “real” en alguna de las dos vistas?
    src_path = None
    for c in (cel_val_ruta, cel_form_ruta):
        if c.hyperlink:
            src_path = c.hyperlink.target
            break                     # listo

    # 2️⃣  Si no, parseamos la fórmula
    if src_path is None:
        formula = cel_form_ruta.value
        if isinstance(formula, str) and formula.startswith('='):
            m = pat.search(formula)
            if m:
                src_path = m.group(1)  # lo que estaba entre las primeras comillas

    if not src_path:
        print(f'⚠️  Ruta no encontrada en fila {cel_val_ruta.row}')
        continue

    src_path = os.path.normpath(src_path)
    print(f'Procesando: {src_path}')

    if not os.path.exists(src_path):
        print(f'   ⚠️  No existe: {src_path}')
        continue

    dest_path = os.path.join(target_dir, os.path.basename(src_path))
    try:
        shutil.move(src_path, dest_path)
        print(f'   ✔️  Movido a {dest_path}')
    except Exception as e:
        print(f'   ❌ Error moviendo {src_path}: {e}')
